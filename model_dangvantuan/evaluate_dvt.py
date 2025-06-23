# ==============================================================
#  evaluate_dvt.py  –  Precision@k & Recall@k cho
#  dangvantuan/vietnamese-document-embedding + Chroma
# ==============================================================

import os, sys, time, ast, json
import pandas as pd
from openpyxl import Workbook, load_workbook
from rich import print as rprint          # hiển thị log đẹp hơn (pip install rich)

# ---------- CẤU HÌNH ----------------------------------------------------------
PERSIST_DIR = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\chroma_db"
GT_PATH     = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\evaluation\Ground_Truth_Full.xlsx"
OUT_XLSX    = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\evaluation\result_dvt.xlsx"
TOP_K       = 3
# ------------------------------------------------------------------------------

# ---------- TÁI SỬ DỤNG HÀM load_vectordb ------------------------------------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(CURRENT_DIR)
from chatbot_dvt import load_vectordb     # → phải tồn tại trong cùng thư mục

# ---------- HÀM TIỆN ÍCH ------------------------------------------------------
def safe_literal_list(cell):
    """
    Nhận '[4,5]' hoặc '4, 5' hoặc '4' → trả về [4,5] hoặc [4].
    Nếu ô trống → [].
    """
    if pd.isna(cell) or str(cell).strip() == "":
        return []
    try:
        # Dùng ast cho trường hợp '[4,5]'
        parsed = ast.literal_eval(str(cell))
        if isinstance(parsed, int):
            return [parsed]
        if isinstance(parsed, list):
            return [int(x) for x in parsed]
    except (ValueError, SyntaxError):
        # Trường hợp '4,5' hoặc '4'
        return [int(x) for x in str(cell).strip(" []").split(",") if x.strip().isdigit()]
    return []

def read_ground_truth(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    # Tìm tên cột bất kể gõ tắt
    col_map = {}
    for col in df.columns:
        normalized = col.lower()
        if "truy" in normalized or "câu" in normalized:
            col_map["query"] = col
        elif "chunk" in normalized:
            col_map["relevant_raw"] = col
        elif "stt" in normalized:
            col_map["STT"] = col
    missing = {"query", "relevant_raw", "STT"} - set(col_map)
    if missing:
        raise ValueError(f"⚠ Không tìm thấy cột bắt buộc: {missing}")

    df = df.rename(columns={
        col_map["query"]: "query",
        col_map["relevant_raw"]: "relevant_raw",
        col_map["STT"]: "STT"
    })

    df["relevant_ids"] = df["relevant_raw"].apply(safe_literal_list)
    df = df[["STT", "query", "relevant_ids"]].copy()

    # --- VALIDATION nhẹ -------------------------------------------------------
    bad_rows = df[df["relevant_ids"].str.len() == 0]
    if not bad_rows.empty:
        bad_list = bad_rows["STT"].tolist()[:10]
        raise ValueError(f"⚠ Có {len(bad_rows)} dòng không có ID ground-truth. "
                         f"Ví dụ STT: {bad_list}")
    return df

def retrieve_ids(retriever, query: str, k: int):
    docs = retriever.get_relevant_documents(query)
    return [int(d.metadata.get("chunk_id")) for d in docs[:k]]

def pr_at_k(retrieved, relevant, k: int):
    if k == 0:
        return 0.0, 0.0
    inter = set(retrieved) & set(relevant)
    precision = len(inter) / k
    recall    = len(inter) / len(relevant) if relevant else 0.0
    return precision, recall

def ensure_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Query", f"P@{TOP_K}", f"R@{TOP_K}",
                   "Elapsed(s)", "Retrieved IDs"])
    return wb, ws

# ---------- CHƯƠNG TRÌNH CHÍNH -----------------------------------------------
def main():
    rprint("[bold cyan]▶ Đang nạp ground-truth …")
    gt = read_ground_truth(GT_PATH)
    rprint(f"   • Tổng truy vấn: [bold]{len(gt)}[/]")

    rprint("[bold cyan]▶ Đang nạp vector database …")
    vectordb = load_vectordb(PERSIST_DIR)
    retriever = vectordb.as_retriever(
        search_type="similarity",
        search_kwargs={"k": TOP_K}
    )
    rprint("   • Retriever ready ✔")

    # Validate thử 3 query đầu
    sample_retrieved = retrieve_ids(retriever, gt.iloc[0]["query"], TOP_K)
    if not isinstance(sample_retrieved, list):
        raise RuntimeError("⚠ retriever.get_relevant_documents() không trả về 'chunk_id'!")

    rprint("[green]✔ Validation OK – Bắt đầu đánh giá[/]")
    wb, ws = ensure_workbook(OUT_XLSX)

    processed = 0
    for _, row in gt.iterrows():
        stt   = int(row["STT"])
        query = row["query"]
        relevant = row["relevant_ids"]

        t0 = time.time()
        retrieved = retrieve_ids(retriever, query, TOP_K)
        elapsed = round(time.time() - t0, 3)

        p, r = pr_at_k(retrieved, relevant, TOP_K)
        ws.append([stt, query, p, r, elapsed, json.dumps(retrieved, ensure_ascii=False)])
        processed += 1

        if processed % 20 == 0:
            wb.save(OUT_XLSX)
            rprint(f"   ▸ Đã xử lý {processed}/{len(gt)} truy vấn …")

    wb.save(OUT_XLSX)

    # ---------- TÓM TẮT -------------------------------------------------------
    df_res = pd.read_excel(OUT_XLSX)
    macro_p = df_res[f"P@{TOP_K}"].mean()
    macro_r = df_res[f"R@{TOP_K}"].mean()

    rprint("\n[bold yellow]== KẾT THÚC ==[/]")
    rprint(f"Số truy vấn:        {processed}")
    rprint(f"Macro Precision@{TOP_K}: [bold]{macro_p:.3f}[/]")
    rprint(f"Macro Recall@{TOP_K}:    [bold]{macro_r:.3f}[/]")
    rprint(f"Kết quả Excel:      [green]{OUT_XLSX}[/]")

if __name__ == "__main__":
    main()
