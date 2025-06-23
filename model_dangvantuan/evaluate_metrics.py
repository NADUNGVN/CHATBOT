# ==============================================================
#  evaluate_metrics.py  –  Tính 6 chỉ số: P@k, R@k, MRR, nDCG@k,
#                          P95 latency & Throughput cho ChromaDB
# ==============================================================

import os, sys, time, ast, json, math
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from rich import print as rprint          # pip install rich

# ---------- CẤU HÌNH ----------------------------------------------------------
PERSIST_DIR = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\chroma_db"
GT_PATH     = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\evaluation\Ground_Truth_Full.xlsx"
OUT_XLSX    = r"E:\WORK\project\chatbot_RAG\model_dangvantuan\evaluation\result_all_metrics.xlsx"
TOP_K       = 3                          
# ------------------------------------------------------------------------------

# ---------- NẠP HÀM load_vectordb --------------------------------------------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(CURRENT_DIR)
from chatbot_dvt import load_vectordb     

# ---------- HÀM TIỆN ÍCH ------------------------------------------------------
def safe_literal_list(cell):
    "Chuyển ô Excel thành list[int] an toàn."
    if pd.isna(cell) or str(cell).strip() == "":
        return []
    try:
        parsed = ast.literal_eval(str(cell))
        if isinstance(parsed, int):
            return [parsed]
        if isinstance(parsed, list):
            return [int(x) for x in parsed]
    except (ValueError, SyntaxError):
        pass
    return [int(x) for x in str(cell).strip(" []").split(",") if x.strip().isdigit()]

def read_ground_truth(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    col_map = {}
    for c in df.columns:
        lc = c.lower()
        if "truy" in lc or "câu" in lc: col_map["query"] = c
        elif "chunk" in lc:             col_map["relevant_raw"] = c
        elif "stt" in lc:               col_map["stt"] = c
    missing = {"query", "relevant_raw", "stt"} - col_map.keys()
    if missing:
        raise ValueError(f"Thiếu cột: {missing}")

    df = df.rename(columns={
        col_map["query"]: "query",
        col_map["relevant_raw"]: "relevant_raw",
        col_map["stt"]: "STT"
    })
    df["relevant_ids"] = df["relevant_raw"].apply(safe_literal_list)
    bad = df[df["relevant_ids"].str.len() == 0]
    if not bad.empty:
        raise ValueError(f"{len(bad)} dòng thiếu ground-truth (VD STT {bad.iloc[0]['STT']})")
    return df[["STT", "query", "relevant_ids"]]

def retrieve_ids(retriever, query: str, k: int):
    docs = retriever.get_relevant_documents(query)
    return [int(d.metadata.get("chunk_id")) for d in docs[:k]]

def precision_recall_at_k(retrieved, relevant, k: int):
    inter = set(retrieved[:k]) & set(relevant)
    p = len(inter) / k
    r = len(inter) / len(relevant) if relevant else 0.0
    return p, r

def reciprocal_rank(retrieved, relevant):
    for idx, doc_id in enumerate(retrieved, 1):
        if doc_id in relevant:
            return 1.0 / idx
    return 0.0

def dcg_at_k(retrieved, relevant, k: int):
    dcg = 0.0
    for idx, doc_id in enumerate(retrieved[:k], 1):
        rel = 1.0 if doc_id in relevant else 0.0
        dcg += rel if idx == 1 else rel / math.log2(idx + 1)
    return dcg

def idcg_at_k(num_rel, k: int):
    idcg = 0.0
    for idx in range(1, min(num_rel, k) + 1):
        idcg += 1.0 if idx == 1 else 1.0 / math.log2(idx + 1)
    return idcg

def ensure_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Query",
                   f"P@{TOP_K}", f"R@{TOP_K}", "RR", f"nDCG@{TOP_K}",
                   "Latency(s)", "Retrieved IDs"])
    return wb, ws

# ---------- CHƯƠNG TRÌNH CHÍNH -----------------------------------------------
def main():
    rprint("[bold cyan]▶ Đang nạp ground-truth …")
    gt = read_ground_truth(GT_PATH)
    rprint(f"   • Tổng truy vấn: [bold]{len(gt)}[/]")

    rprint("[bold cyan]▶ Đang nạp vector DB …")
    vectordb = load_vectordb(PERSIST_DIR)
    retriever = vectordb.as_retriever(
        search_type="similarity",
        search_kwargs={"k": TOP_K}
    )
    rprint("[green]✔ DB sẵn sàng. Bắt đầu đánh giá[/]")

    wb, ws = ensure_workbook(OUT_XLSX)

    # ----- Biến tích lũy ------------------------------------------------------
    elapsed_times = []
    p_sum = r_sum = rr_sum = ndcg_sum = 0.0
    processed = 0

    # ----- Vòng lặp đánh giá --------------------------------------------------
    for _, row in gt.iterrows():
        stt, query, relevant = int(row["STT"]), row["query"], row["relevant_ids"]

        t0 = time.time()
        retrieved = retrieve_ids(retriever, query, TOP_K)
        latency = time.time() - t0
        elapsed_times.append(latency)

        p, r = precision_recall_at_k(retrieved, relevant, TOP_K)
        rr = reciprocal_rank(retrieved, relevant)
        dcg = dcg_at_k(retrieved, relevant, TOP_K)
        idcg = idcg_at_k(len(relevant), TOP_K)
        ndcg = dcg / idcg if idcg > 0 else 0.0

        # lưu vào excel
        ws.append([stt, query, p, r, rr, ndcg, round(latency, 3),
                   json.dumps(retrieved, ensure_ascii=False)])

        # tích lũy
        p_sum   += p
        r_sum   += r
        rr_sum  += rr
        ndcg_sum += ndcg
        processed += 1

        if processed % 20 == 0:
            wb.save(OUT_XLSX)
            rprint(f"   ▸ Đã xử lý {processed}/{len(gt)} …")

    wb.save(OUT_XLSX)

    # ---------- TÓM TẮT KẾT QUẢ ---------------------------------------------
    N = processed
    macro_p   = p_sum   / N
    macro_r   = r_sum   / N
    mrr       = rr_sum  / N
    mean_ndcg = ndcg_sum / N
    p95_lat   = np.percentile(np.array(elapsed_times), 95)
    qps       = N / sum(elapsed_times)

    rprint("\n[bold yellow]== KẾT THÚC ==[/]")
    rprint(f"Số truy vấn:               {N}")
    rprint(f"Macro Precision@{TOP_K}:   [bold]{macro_p:.3f}[/]")
    rprint(f"Macro Recall@{TOP_K}:      [bold]{macro_r:.3f}[/]")
    rprint(f"MRR:                       [bold]{mrr:.3f}[/]")
    rprint(f"Mean nDCG@{TOP_K}:         [bold]{mean_ndcg:.3f}[/]")
    rprint(f"P95 latency (s):           [bold]{p95_lat:.3f}[/]")
    rprint(f"Throughput (QPS):          [bold]{qps:.2f}[/]")
    rprint(f"Kết quả chi tiết Excel:    [green]{OUT_XLSX}[/]")

if __name__ == "__main__":
    main()
